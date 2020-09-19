from openpyxl import Workbook
from openpyxl import load_workbook
import slot
import json


def getkey(a, b):
    '''returns key for cell[a][b] in excel'''
    return str(chr(ord('A')+a-1))+str(b)

# copy same values in merged cells
def sanitise_sheets(valid_sheets, wb):
    for sheet in valid_sheets:
        ws = wb.get_sheet_by_name(sheet)
        ws.delete_cols(1,2)
        ws.delete_rows(1,2)


def format_cell(original_value):
    '''converts cell text to actually usable array of subject name and room numbers'''
    if(original_value is None):
        pass
    elif("physics" in original_value.lower()):
        original_value = "phy_lab"
    elif("drawing" in original_value.lower()):
        original_value = "ed_lab"
    elif("data" in original_value.lower()):
        original_value = "pds_lab"
    elif("electrical" in original_value.lower()):
        original_value = "et_lab"
    elif("process" in original_value.lower()):
        original_value = "mech_lab"
    elif("language" in original_value.lower()):
        original_value = "hs_lab"
    elif("chemistry" in original_value.lower()):
        original_value = "chem_lab"
    elif("l u n c h" in original_value.lower()):
        original_value = "lunch_break"
    return [x for sub in original_value.split("\n") for j in sub.split(" ") for x in j.split(",") if x]

def match_slots(l1, l2):
    if(len(l1) > len(l2)):
        return False
    count = 0
    for i in range(0, len(l1)):
        if(l1[i] in l2):
            count += 1
    return count

def map_sheets(valid_sheets, wb):
    slot_fillup = set()

    for sheet in valid_sheets:
        sub_vs_room = {}
        sub_vs_slot = {}
        ws = wb.get_sheet_by_name(sheet)
        if(ws['L2'].value is not None or ws['K2'].value is None):
            print("parsing error in sheet {}, please inclue it in first_year.csv manually".format(sheet))
            continue

        for i in range(0, 6):
            for j in range(0, 10):
                if j == 5:
                    continue
                cell = ws[getkey(j+2, i+3)].value
                if(cell is not None):
                    cell_value = format_cell(cell)
                    subject = cell_value[0];
                    sub_vs_room[subject] = cell_value[1:]
                    if(subject not in sub_vs_slot):
                        sub_vs_slot[subject] = list()
                    sub_vs_slot[subject].append(str(i)+str(j-1 if j>5 else j))

        sub_to_slot_name = {}
        for subject in sub_vs_slot:
            max_match = 0
            for slot_name in slot.slot_map:
                present_match = match_slots(sub_vs_slot[subject], slot.slot_map[slot_name])
                if(present_match > max_match):
                    sub_to_slot_name[subject] = slot_name
                    max_match = present_match

        for subject in sub_to_slot_name:
            if('(' in subject or not sub_vs_room[subject]):
                continue
            for room in sub_vs_room[subject]:
                slot_fillup.add( (subject, sub_to_slot_name[subject], room) )

    return list(slot_fillup)


def format_excel(filename):

    wb = load_workbook(filename=filename)
    sheets = wb.sheetnames
    valid_sheets = [sheet for sheet in sheets if (wb.get_sheet_by_name(sheet)['D10'].value == "EAA" or wb.get_sheet_by_name(sheet)['E10'].value == "EAA")]

    print("count of sheets with timetable : {}".format(len(valid_sheets)))

    sanitise_sheets(valid_sheets, wb)
    wb.save(filename.split('.')[0]+"_new.xlsx")

    slot_fillup = map_sheets(valid_sheets, wb)

    slot_fillup = sorted(slot_fillup)
    return slot_fillup


if __name__ == "__main__":
    filename = "wb.xlsx"
    slot_fillup = format_excel(filename)
    print("size of slot fillups: ", len(slot_fillup))
    open("slot_fillup.json", "w").write(json.dumps(slot_fillup))