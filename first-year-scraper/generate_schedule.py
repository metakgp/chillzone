from openpyxl import load_workbook
import generate_subjectDetails
import json



def getkey(a, b):
    '''returns key for cell[a][b] in excel'''
    return str(chr(ord('A')+a-1))+str(b)

def format_cell(original_value):
    '''converts cell text to actually usable array of subject name and room numbers'''
    if(original_value is None):
        pass
    return [x for sub in original_value.split("\n") for j in sub.split(" ") for x in j.split(",") if x]

def generate_schedule(valid_sheets, workbook, subjects_dict):
    with open("../frontend/src/schedule.json", "r") as json_data:
        schedule_dict = json.load(json_data)
    with open("../frontend/src/empty_schedule.json", "r") as json_data:
        empty_schedule_dict = json.load(json_data)

    for sheet in valid_sheets:
        worksheet = workbook[sheet]

        if(worksheet['N4'].value is not None or worksheet['M4'].value is None):
            print("parsing error in sheet {}".format(sheet))

            # for aut2023.pdf only page 5 has parsing error in which one column is repeated.
            # manually deleting that column:
            worksheet.delete_cols(7,1)
        
        worksheet.delete_cols(9,1)
        for i in range(0, 6):
            for j in range(0, 9):
                cell = worksheet[getkey(j+4, i+5)].value
                if cell is not None and ("NR" in cell or "NC" in cell):
                    cell_value = format_cell(cell)
                    for room in cell_value[1::]:
                        if(schedule_dict[room][i][j]==""):
                            schedule_dict[room][i][j] = subjects_dict[cell_value[0]][0]
                        if(room in empty_schedule_dict[i][j]):
                            empty_schedule_dict[i][j].remove(room)
                                
    with open("../frontend/src/schedule.json", "w") as outfile:
        json.dump(schedule_dict, outfile, indent=2)

    with open("../frontend/src/empty_schedule.json", "w") as outfile:
        json.dump(empty_schedule_dict, outfile, indent=2)


def format_excel(filename, subjects_dict):
    workbook = load_workbook(filename=filename)
    sheets = workbook.sheetnames
    valid_sheets = [sheet for sheet in sheets if (workbook[sheet]['D10'].value == "EAA" or workbook[sheet]['E10'].value == "EAA")]

    print("count of sheets with timetable : {}".format(len(valid_sheets)))

    generate_schedule(valid_sheets, workbook, subjects_dict)


if __name__=="__main__":
    format_excel("workbook.xlsx", generate_subjectDetails.generate_subjectDetails())    
