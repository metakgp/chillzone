import camelot
from openpyxl import load_workbook
import json


def parse_pdf(filename, target_filename):
    print("Started processing pdf file. This might take a while...")
    tables = camelot.read_pdf(filename, pages='1-end', copy_text=['h']) # if you are here to change args please be sure that you know what you are doing
    tables.export(target_filename, f="excel")
    print("Done parsing pdf. exported as: ", target_filename)

def sanitise_sheets(valid_sheets, wb):
    for sheet in valid_sheets:
        ws = wb.get_sheet_by_name(sheet)
        ws.delete_cols(1,2)
        ws.delete_rows(1,2)

def getkey(a, b):
    '''returns key for cell[a][b] in excel'''
    return str(chr(ord('A')+a-1))+str(b)

def format_cell(original_value):
    '''converts cell text to actually usable array of subject name and room numbers'''
    if(original_value is None):
        pass
    return [x for sub in original_value.split("\n") for j in sub.split(" ") for x in j.split(",") if x]

def map_sheets(valid_sheets, wb):
    slot_fillup = set()

    for sheet in valid_sheets:
        sub_vs_room = {}
        sub_vs_slot = {}
        ws = wb.get_sheet_by_name(sheet)

        if(ws['L2'].value is not None or ws['K2'].value is None):
            print("parsing error in sheet {}, please include it in first_year.csv manually".format(sheet))
            continue

        for i in range(0, 6):
            for j in range(0, 10):
                if j == 5:
                    continue
                cell = ws[getkey(j+2, i+3)].value
                if(cell is not None):
                    cell_value = format_cell(cell)

def format_excel(filename):

    wb = load_workbook(filename=filename)
    sheets = wb.sheetnames
    valid_sheets = [sheet for sheet in sheets if (wb.get_sheet_by_name(sheet)['D10'].value == "EAA" or wb.get_sheet_by_name(sheet)['E10'].value == "EAA")]

    print("count of sheets with timetable : {}".format(len(valid_sheets)))

    sanitise_sheets(valid_sheets, wb)
    wb.save(filename.split('.')[0]+"_new.xlsx")

    map_sheets(sheets, wb)

if __name__ == "__main__":
    parse_pdf("spr2019.pdf","test.xlsx")