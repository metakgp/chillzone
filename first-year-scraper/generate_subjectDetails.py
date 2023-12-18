from openpyxl import load_workbook
import json

def get_cell_value(a, b):
    '''returns cell value for cell[a][b] in excel'''
    return sheet[str(chr(ord('A')+a-1))+str(b)].value

def generate_subjectDetails():

    workbook = load_workbook("workbook.xlsx")
    global sheet
    sheet = workbook[workbook.sheetnames[2]]

    subjects_dict = {}

    for i in range(0, 15):
        if not ("Laboratory" in get_cell_value(5, i+4)):
            if(get_cell_value(6, i+4)[2:3] == "0"):
                subjects_dict[get_cell_value(4, i+4)[0:2]] = [
                    get_cell_value(4, i+4),
                    get_cell_value(5, i+4).upper(),
                ]
            else:
                subjects_dict[get_cell_value(4, i+4)[0:2]] = [
                    get_cell_value(4, i+4),
                    get_cell_value(5, i+4).upper(),
                ]
                subjects_dict[get_cell_value(4, i+4)[0:2]+"(T)"] = [
                    get_cell_value(4, i+4)+"(T)",
                    get_cell_value(5, i+4).upper()+" TUTORIAL",
                ]

    with open("../frontend/src/data/subjectDetails.json", "r") as json_data:
        subject_details = json.load(json_data)
    
    for key in subjects_dict.keys():
        subject_details[subjects_dict[key][0]] = subjects_dict[key][1]

    with open("../frontend/src/data/subjectDetails.json", "w") as outfile:
        json.dump(subject_details, outfile, indent=2)

    return subjects_dict

    
if __name__=="__main__":
    generate_subjectDetails()    
